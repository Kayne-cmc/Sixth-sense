import openpyxl
import os
import time

from constants import TEST_TITLES


class Benchmark:
    def __init__(self, filename):
        self.sheets = {}
        self.start_time = None
        self.end_time = None
        self.filename = filename

        if os.path.isfile(filename):
            self.workbook = openpyxl.load_workbook(filename)
        else:
            self.workbook = openpyxl.Workbook()
        
        for _ , title in TEST_TITLES.items():
            if title not in self.workbook.sheetnames:
                self.sheets[title] = self.workbook.create_sheet(title)

    def start_test(self):
        self.start_time = time.time()
    
    def end_test(self, sheet_title):
        if self.start_time is None or self.sheets[sheet_title] is None:
            return
        
        self.end_time = time.time()
        time_difference =  self.end_time - self.start_time
        empty_row = self.sheet[sheet_title].max_row + 1
        self.sheets[sheet_title].cell(row = empty_row, column = 1, value=self.start_time)
        self.sheets[sheet_title].cell(row = empty_row, column = 2, value=self.end_time)
        self.sheets[sheet_title].cell(row = empty_row, column = 3, value=time_difference)

        self.start_time = None
        self.end_time = None

    def save_benchmark(self):
        self.workbook.save(self.filename)